import streamlit as st
import csv
from collections import defaultdict
from datetime import datetime
import io
import os

# Try to import openpyxl for nice Excel formatting
try:
    import openpyxl
    from openpyxl.styles import Font
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    st.warning("openpyxl not installed. Basic CSV download will still work.")

st.set_page_config(page_title="Journal Account Summary", layout="centered")
st.title("📊 Journal Account Summary Tool")
st.markdown("Upload your journal CSV file → Get a clean account summary Excel file (named by journal date)")

# File uploader
uploaded_file = st.file_uploader("Choose your Journal CSV file", type=["csv"])

if uploaded_file is not None:
    # Read the file content
    try:
        content = uploaded_file.getvalue().decode('utf-8')
        lines = content.splitlines()
        reader = csv.DictReader(lines)
        
        required = {'Account Name', 'Debits', 'Credits', 'Journal Date'}
        if not required.issubset(set(reader.fieldnames or [])):
            st.error("Missing required columns. Please make sure your CSV has: Account Name, Debits, Credits, Journal Date")
            st.stop()
        
        # Get Journal Date from first row
        first_row = next(reader, None)
        journal_date_str = first_row.get('Journal Date', '').strip() if first_row else None
        
        # Reset reader
        reader = csv.DictReader(lines)
        
        balances = defaultdict(lambda: {'Debits': 0.0, 'Credits': 0.0})
        
        for row in reader:
            account = row.get('Account Name', '').strip()
            if not account:
                continue
            try:
                debits = float(row.get('Debits', 0) or 0)
            except:
                debits = 0.0
            try:
                credits = float(row.get('Credits', 0) or 0)
            except:
                credits = 0.0
            
            balances[account]['Debits'] += debits
            balances[account]['Credits'] += credits
        
        # Prepare results
        results = []
        for account in sorted(balances.keys()):
            results.append({
                'Account Name': account,
                'Total Debits': round(balances[account]['Debits'], 2),
                'Total Credits': round(balances[account]['Credits'], 2)
            })
        
        # Create nice output filename
        if journal_date_str:
            try:
                parsed = datetime.strptime(journal_date_str, '%m/%d/%Y')
                date_part = parsed.strftime('%Y-%m-%d')
            except:
                date_part = "Unknown_Date"
        else:
            date_part = "Unknown_Date"
        
        output_filename = f"{date_part}_Account_Summary.xlsx"
        
        # Show preview
        st.success(f"Processed {len(results)} unique accounts. Journal Date: {journal_date_str or 'Not found'}")
        
        # Display table
        st.subheader("Account Summary Preview")
        preview_df = {
            'Account Name': [r['Account Name'] for r in results],
            'Total Debits': [r['Total Debits'] for r in results],
            'Total Credits': [r['Total Credits'] for r in results]
        }
        st.dataframe(preview_df, use_container_width=True)
        
        # Grand totals
        total_debits = sum(r['Total Debits'] for r in results)
        total_credits = sum(r['Total Credits'] for r in results)
        st.info(f"**Grand Total** — Debits: ${total_debits:,.2f} | Credits: ${total_credits:,.2f}")
        
        # Download button
        if EXCEL_AVAILABLE:
            # Create Excel in memory
            output = io.BytesIO()
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Account Summary"
            
            ws['A1'] = "Account Summary"
            ws['A1'].font = Font(bold=True, size=14)
            
            ws['A3'] = "Account Name"
            ws['B3'] = "Total Debits"
            ws['C3'] = "Total Credits"
            for cell in ['A3', 'B3', 'C3']:
                ws[cell].font = Font(bold=True)
            
            for i, res in enumerate(results, start=4):
                ws.cell(row=i, column=1, value=res['Account Name'])
                ws.cell(row=i, column=2, value=res['Total Debits'])
                ws.cell(row=i, column=3, value=res['Total Credits'])
            
            total_row = len(results) + 5
            ws.cell(row=total_row, column=1, value="GRAND TOTAL").font = Font(bold=True)
            ws.cell(row=total_row, column=2, value=total_debits).font = Font(bold=True)
            ws.cell(row=total_row, column=3, value=total_credits).font = Font(bold=True)
            
            ws.column_dimensions['A'].width = 45
            ws.column_dimensions['B'].width = 18
            ws.column_dimensions['C'].width = 18
            
            wb.save(output)
            output.seek(0)
            
            st.download_button(
                label="📥 Download Excel Summary",
                data=output,
                file_name=output_filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        else:
            st.info("For better Excel formatting, run `pip install openpyxl` locally.")
            # Fallback CSV download
            import pandas as pd
            df = pd.DataFrame(results)
            csv_data = df.to_csv(index=False).encode('utf-8')
            st.download_button(
                label="📥 Download as CSV",
                data=csv_data,
                file_name=output_filename.replace('.xlsx', '.csv'),
                mime="text/csv"
            )
            
    except Exception as e:
        st.error(f"Error processing file: {str(e)}")

st.caption("Free web app • Upload your journal CSV • Automatically named by Journal Date")